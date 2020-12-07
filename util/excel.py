# coding: utf-8

from openpyxl import load_workbook, Workbook


class Excel:
    def __init__(self, filename=None, write=False):
        self.filename = filename
        if write is False:
            self.wb = load_workbook(filename=filename)
        else:
            self.wb = Workbook()
        """
        get_sheet_by_name、get_sheet_names已废弃，新写法参见：
        https://stackoverflow.com/questions/51202874/deprecationwarning-call-to-deprecated-function-get-sheet-by-name-use-wbsheetn
        """
        # self.sheet1 = self.wb.get_sheet_by_name(self.wb.get_sheet_names()[0])
        self.sheet1 = self.wb[self.wb.sheetnames[0]]

    def rows(self):
        """
        Usage:
        for row in rows():
            for cell in row:
                print(cell.value)
        """
        return self.sheet1.rows

    def write(self, rows):
        """
        write to excel file
        :param rows: [(row1),(row2)...]
        :return: None
        """
        for row in rows:
            self.sheet1.append(row)

    def save(self):
        self.wb.save(self.filename)